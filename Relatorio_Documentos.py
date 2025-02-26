import os
import subprocess
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def get_diretorio_dados(diretorios_dados, contrato):
    return diretorios_dados.get(contrato)

def get_diretorio_base(diretorios_base, contrato):
    return diretorios_base.get(contrato)

def get_diretorio_funcionario(diretorio_base, nome):
    return os.path.join(os.path.expanduser("~"), diretorio_base, nome)

def formatar_cpf(cpf):
    cpf = ''.join(filter(str.isdigit, str(cpf)))
    cpf = cpf.zfill(11)
    return f"{cpf[:3]}.{cpf[3:6]}.{cpf[6:9]}-{cpf[9:]}"

def formatar_data(data):
    return datetime.strftime(data, '%d/%m/%Y')

def obter_documentos_requeridos(funcao, documentos_por_funcao):
    return documentos_por_funcao.get(funcao, documentos_por_funcao["OUTRAS"])

def gerar_dados_planilha(diretorio_base, tabela_dados, documentos_por_funcao, ordem_documentos):
    dados_planilha = []
    
    for nome, linha in tabela_dados.groupby("NOME"):
        caminho_funcionario = get_diretorio_funcionario(diretorio_base, nome)
        if os.path.isdir(caminho_funcionario):
            funcao = str(linha["DESC FUNÇÃO"].iloc[0])
            admissao = formatar_data(linha["DATA ADMISSAO"].iloc[0])
            cpf = formatar_cpf(linha["CPF"].iloc[0])

            documentos_requeridos = obter_documentos_requeridos(funcao, documentos_por_funcao)
            documentos_na_pasta = os.listdir(caminho_funcionario)
            linha_dados = [nome, funcao, cpf, admissao]
            documentos_pendentes = []

            for documento in ordem_documentos:
                if documento in documentos_requeridos:
                    nome_esperado = f"{documento} - {nome}.pdf"
                    linha_dados.append("OK" if nome_esperado in documentos_na_pasta else "P")

                    if nome_esperado not in documentos_na_pasta:
                        documentos_pendentes.append(documento)
                else:
                    linha_dados.append("N/A")
            
            linha_dados.append(" - ".join(documentos_pendentes) if documentos_pendentes else "---")
            dados_planilha.append(linha_dados)
    return dados_planilha

def personalizar_planilha(ws, caminho_logo):
    fundo_azul = PatternFill(start_color="003399", end_color="003399", fill_type="solid")
    fonte_branca = Font(size=12, color="FFFFFF", bold=True)
    verde_claro = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  
    vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    alinhamento_esquerda = Alignment(horizontal="left", vertical="center", wrap_text=False, shrink_to_fit=True)
    alinhamento_central = Alignment(horizontal="center", vertical="center", wrap_text=False, shrink_to_fit=True)

    ws.insert_rows(1)

    img = Image(caminho_logo)
    img.width = 250
    img.height = 70
    ws.add_image(img, "A1")

    ws["A1"] = "CONTROLE DE DOCUMENTAÇÃO FUNCIONÁRIOS"
    ws["A1"].font = Font(bold=True, size=22, color="003399")
    ws["A1"].alignment = alinhamento_central

    ws.row_dimensions[1].height = 70
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column) 

    borda = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    ws.row_dimensions[2].height = 34

    for cell in ws[2]:
        cell.fill = fundo_azul
        cell.font = fonte_branca
        cell.alignment = alinhamento_central
        cell.border = borda

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 23
        for idx, cell in enumerate(row):
            cell.border = borda
            if idx == 0:  
                cell.alignment = alinhamento_esquerda

            else:  
                cell.alignment = alinhamento_central

            if cell.value == "OK":
                cell.fill = verde_claro

            elif cell.value == "P":
                cell.fill = vermelho

    ajustar_largura_colunas(ws)

    ws.freeze_panes = "D3"
    ws.auto_filter.ref = "A2:O{}".format(ws.max_row)

def ajustar_largura_colunas(ws):
    tamanho_minimo_documentos_pendentes = 28
    tamanho_minimo_documentos = 8
    tamanho_minimo_nome = 25

    for idx, column_cells in enumerate(ws.columns):
        max_length = max(len(str(cell.value)) for cell in column_cells if cell.value)
        valid_cells = [cell for cell in column_cells if not isinstance(cell, MergedCell) and cell.value]
        if not valid_cells:
            continue
        column = valid_cells[0].column_letter

        if idx == 0:  
            ws.column_dimensions[column].width = max(max_length + 2, tamanho_minimo_nome)

        elif idx == 15: 
            ws.column_dimensions[column].width = max(max_length + 2, tamanho_minimo_documentos_pendentes)

        elif idx >= 4:  
            ws.column_dimensions[column].width = max(max_length + 2, tamanho_minimo_documentos)

        else:
            ws.column_dimensions[column].width = max_length + 2

def gerar_relatorio(contratos, caminho_logo, documentos_por_função, diretorios_base, diretorios_dados):
    wb = Workbook()
    ordem_documentos = ["ASO", "FRE", "EPI", "NR6", "NR10", "NR11", "NR12", "NR18", "NR33", "NR35", "OS"]
    
    for contrato in contratos:
        diretorio_dados = get_diretorio_dados(diretorios_dados, contrato)
        diretorio_base = get_diretorio_base(diretorios_base, contrato)
        tabela_dados = pd.read_excel(diretorio_dados)

        dados_planilha = gerar_dados_planilha(diretorio_base, tabela_dados, documentos_por_função, ordem_documentos)
        colunas = ["FUNCIONÁRIO", "FUNÇÃO", "CPF", "ADMISSÃO"] + ordem_documentos + ["DOCUMENTAÇÃO PENDENTE"]
        df = pd.DataFrame(dados_planilha, columns=colunas)

        ws = wb.create_sheet(title=contrato)  
        for r_idx, row in enumerate([colunas] + df.values.tolist(), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        personalizar_planilha(ws,caminho_logo)
    
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    data_atual = datetime.now().strftime("%d-%m-%Y")
    caminho_saida =  f"RELATÓRIO_DOCUMENTAÇÃO {data_atual}.xlsx"
    wb.save(caminho_saida)
    subprocess.run(["cmd", "/c", "start", "", caminho_saida], shell=True)


# Ponto de Alteração
documentos_por_função = {

}
diretorios_funcionarios = {
    
}
diretorios_dados = {
   
}
contratos = ["OB186 - INHAÚMA", "OB201 - SÃO GONÇALO"]
caminho_logo = "LOGO.png")

gerar_relatorio(contratos, caminho_logo, documentos_por_função, diretorios_base, diretorios_dados)

















               

                    




