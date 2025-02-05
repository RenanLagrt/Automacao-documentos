import os
import subprocess
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def get_diretorio_funcionario(nome):
    return os.path.join(
        os.path.expanduser("~"),
        "CONSORCIO CONCREJATOEFFICO LOTE 1",
        "Central de Arquivos - QSMS",
        "000 ATUAL - OBRA 186 - INHAÚMA",
        "Documentação Funcionários",
        nome
    )

def formatar_cpf(cpf):
    cpf = ''.join(filter(str.isdigit, str(cpf)))
    cpf = cpf.zfill(11)
    return f"{cpf[:3]}.{cpf[3:6]}.{cpf[6:9]}-{cpf[9:]}"

def formatar_data(data):
    return datetime.strftime(data, '%d/%m/%Y')

def obter_documentos_requeridos(funcao, documentos_por_função):
    return documentos_por_função.get(funcao, documentos_por_função["OUTRAS"])

def gerar_dados_planilha(tabela_dados, documentos_por_função, ordem_documentos):
    dados_planilha = []

    for nome, linha in tabela_dados.groupby("NOME"):
        caminho_funcionario = get_diretorio_funcionario(nome)

        if os.path.isdir(caminho_funcionario):
            if not linha.empty:
                funcao = str(linha["DESC FUNÇÃO"].iloc[0])
                admissao = formatar_data(linha["DATA ADMISSAO"].iloc[0])
                cpf = formatar_cpf(linha["CPF"].iloc[0])

                documentos_requeridos = obter_documentos_requeridos(funcao, documentos_por_função)
                documentos_na_pasta = os.listdir(caminho_funcionario)
                linha_dados = [nome, funcao, cpf, admissao]
                documentos_pendentes = []

                for documento in ordem_documentos:
                    if documento in documentos_requeridos:
                        nome_esperado = f"{documento} - {nome}.pdf"

                        if nome_esperado in documentos_na_pasta:
                            linha_dados.append("OK")
                        else:
                            linha_dados.append("Pendente")
                            documentos_pendentes.append(documento)
                    else:
                        linha_dados.append("---")

                linha_dados.append(" - ".join(documentos_pendentes) if documentos_pendentes else "---")
                dados_planilha.append(linha_dados)

    return dados_planilha

def salvar_excel(df, caminho_saida):
    df.to_excel(caminho_saida, index=False)
    wb = load_workbook(caminho_saida)
    ws = wb.active
    personalizar_planilha(ws)
    wb.save(caminho_saida)

def personalizar_planilha(ws):
    fundo_preto = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    fonte_branca = Font(color="FFFFFF", bold=True)

    alinhamento_esquerda = Alignment(horizontal="left", vertical="center", wrap_text=False, shrink_to_fit=True)
    alinhamento_central = Alignment(horizontal="center", vertical="center", wrap_text=False, shrink_to_fit=True)

    borda = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Aplicar estilo ao cabeçalho
    for cell in ws[1]:
        cell.fill = fundo_preto
        cell.font = fonte_branca
        cell.alignment = alinhamento_central

    # Aplicar alinhamento e bordas para as células de dados
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for idx, cell in enumerate(row):
            cell.border = borda
            if idx == 0:  # Nome do funcionário (alinhado à esquerda)
                cell.alignment = alinhamento_esquerda
            else:  # Demais colunas (alinhamento central)
                cell.alignment = alinhamento_central

    ajustar_largura_colunas(ws)

    ws.freeze_panes = ws["E2"]
    ws.auto_filter.ref = "A1:B{}".format(ws.max_row)

def ajustar_largura_colunas(ws):
    tamanho_minimo_documentos_pendentes = 28
    tamanho_minimo_documentos = 8
    tamanho_minimo_nome = 25

    for idx, column_cells in enumerate(ws.columns):
        max_length = max(len(str(cell.value)) for cell in column_cells if cell.value)
        column = column_cells[0].column_letter

        if idx == 0:  # Primeira coluna (Nome)
            ws.column_dimensions[column].width = max(max_length + 2, tamanho_minimo_nome)

        elif idx == 15:  # Coluna de Documentos Pendentes
            ws.column_dimensions[column].width = max(max_length + 2, tamanho_minimo_documentos_pendentes)

        elif idx >= 4:  # Demais colunas de documentos
            ws.column_dimensions[column].width = max(max_length + 2, tamanho_minimo_documentos)

        else:
            ws.column_dimensions[column].width = max_length + 2

def gerar_relatorio():
    diretorio_dados = os.path.join(os.path.expanduser("~"), "CONSORCIO CONCREJATOEFFICO LOTE 1", "Central de Arquivos - QSMS", "000 ATUAL - OBRA 186 - INHAÚMA", "Efetivo", "QUANTITATIVO COMPARTILHAR.xlsx")
    tabela_dados = pd.read_excel(diretorio_dados)

    documentos_por_função = {
    "ELETRICISTA DE REPARO DE REDE DE SANEAMENTO" : [
        "ASO", "FRE", "FICHA EPI", "NR6", "NR10", "NR12", "NR18", "NR33", "NR35", "OS"
    ],
    "OPERADOR DE REPARO DE REDE DE SANEAMENTO" : [
        "ASO", "FRE", "FICHA EPI", "NR6", "NR12", "NR18", "NR33", "NR35", "OS"
    ],
    "1/2 OFICIAL DE REPARO DE REDE DE SANEAMENTO CIVIL" : [
        "ASO", "FRE", "FICHA EPI", "NR6", "NR12", "NR18", "NR33", "NR35", "OS"
    ],
    "AUXILIAR DE REPARO DE REDE DE SANEAMENTO" : [
        "ASO", "FRE", "FICHA EPI", "NR6", "NR12", "NR18", "NR33", "NR35", "OS"
    ],
    "ENCARREGADO DE REPARO DE REDE DE SANEAMENTO" : [
        "ASO", "FRE", "FICHA EPI", "NR6", "NR18", "NR33", "NR35", "OS"
    ],
    "OPERADOR RETROESCAVADEIRA" : [
        "ASO", "FRE", "FICHA EPI", "NR6", "NR11", "NR18", "OS"
    ],
    "ESTAGIARIO" : [
        "ASO", "FICHA EPI", "NR6", "NR18", "OS"
    ],
    "OUTRAS": [
        "ASO", "FRE", "FICHA EPI", "NR6", "NR18", "OS"
    ]
}
    ordem_documentos = ["ASO", "FRE", "FICHA EPI", "NR6", "NR10", "NR11", "NR12", "NR18", "NR33", "NR35", "OS"]
    
    dados_planilha = gerar_dados_planilha(tabela_dados, documentos_por_função, ordem_documentos)

    colunas = ["FUNCIONÁRIO", "FUNÇÃO", "CPF", "ADMISSÃO"] + ordem_documentos + ["DOCUMENTAÇÃO PENDENTE"]
    df = pd.DataFrame(dados_planilha, columns=colunas)

    data_atual = datetime.now().strftime("%d-%m-%Y")
    caminho_saida = os.path.join(os.path.expanduser("~"), "CONSORCIO CONCREJATOEFFICO LOTE 1", "Central de Arquivos - QSMS", "000 ATUAL - OBRA 186 - INHAÚMA", "Documentação Funcionários",f"RELATÓRIO_DOCUMENTAÇÃO_INHAUMA {data_atual}.xlsx")

    salvar_excel(df, caminho_saida)
    subprocess.run(["cmd", "/c", "start", "", caminho_saida], shell=True)

gerar_relatorio()